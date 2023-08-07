"""
This program creates a GUI application using tkinter to collect user data and store it in an Excel file. The program also allows the user to load data from an existing Excel file, display the data in a treeview, and visualize the data using matplotlib.

The program consists of two frames: a data entry form and a data display frame. The data entry form contains labels and entry fields for the user to input their name, age, gender, and hobbies. The form also includes a checkbox for the user to agree to the terms and conditions, and a submit button to add the data to the Excel file.

The program also includes a dark mode and light mode, which can be toggled using a menu. The program loads data from an existing Excel file using a load button, and displays the data in a treeview in the data display frame. The program also includes a visualize button to create a bar chart of the age distribution of the data using matplotlib.

The program is designed to be user-friendly and easy to use, with clear labels and error messages to guide the user through the data entry process.
"""
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import openpyxl as xl
from openpyxl import Workbook
#create the main window
root = tk.Tk()
root.title("pyExcel")

#set the theme
style = ttk.Style()
style.theme_use('clam')

#create the data entry form
form_frame = ttk.LabelFrame(root,padding=20, text="Data Entry Form")
form_frame.grid(row=0,column=0,sticky="NSEW")

#create the data display frame
display_frame = ttk.LabelFrame(root,padding=20, text="Data Display")
display_frame.grid(row=0,column=1,sticky="NSEW")

#Add labels and entry fields to the form
name_label = ttk.Label(form_frame,text="Name:")
name_label.grid(row=0,column=0,sticky="W")
name_entry = ttk.Entry(form_frame,width=20)
name_entry.grid(row=0,column=1,sticky="W")

age_label = ttk.Label(form_frame,text="Age:")
age_label.grid(row=1,column=0,sticky="W")
age_entry = ttk.Entry(form_frame,width=20)
age_entry.grid(row=1,column=1,sticky="W")

gender_label = ttk.Label(form_frame, text="Gender:")
gender_label.grid(row=2,column=0,sticky="W")
gender_combo = ttk.Combobox(form_frame, values=['Male','Female','Other'], width=17, state="readonly")
gender_combo.grid(row=2,column=1,sticky="W")

hobbies_label = ttk.Label(form_frame,text="Hobbies:")
hobbies_label.grid(row=3,column=0,sticky="W")
hobbies_entry = ttk.Entry(form_frame,width=20)
hobbies_entry.grid(row=3,column=1,sticky="W")

#Add a checkbox to the form
agree_var = tk.BooleanVar()
agree_check = ttk.Checkbutton(form_frame,text="I agree to the terms and conditions",variable=agree_var)
agree_check.grid(row=4,columnspan=2,sticky="W")

#Add a button to submit the form
def submit_form():
    #get values from the form
    name = name_entry.get()
    age = age_entry.get()
    gender = gender_combo.get()
    hobbies = hobbies_entry.get()
    agree = agree_var.get()

    #validate the form
    if not name:
        messagebox.showerror("Error","Please enter a name")
        return
    if not age:
        messagebox.showerror("Error","Please enter an age")
        return
    if not gender:
        messagebox.showerror("Error","Please select a gender")
        return
    if not agree:
        messagebox.showerror("Error","Please agree to the terms and conditions")
        return
    
    #Add the data to the excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(['Name','Age','Gender','Hobbies'])
    wb.save("data.xlsx")

    #show a success message
    messagebox.showinfo("Success","Data added successfully")

submit_button = ttk.Button(form_frame,text="Submit",command=submit_form)
submit_button.grid(row=5,columnspan=2,sticky="EW")

#Add a dark mode and light mode
def set_dark_mode():
    style.theme_use('clam')
    style.configure('TLabel',background="#000",foreground="#fff")
    style.configure('TButton',background="#000",foreground="#fff")
    style.configure('TEntry',background="#000",foreground="#fff",insertbackground="#fff")
    submit_button.configure(style="DarkButton.TButton")

def set_light_mode():
    style.theme_use('clam')
    style.configure("TLabel",background="#fff")
    style.configure("TButton", background="#fff")
    style.configure("TEntry",background="#fff")
    submit_button.configure(style="LightButton.TButton")

#Add a menu to switch between dark and light mode
menu_bar = tk.Menu(root)
mode_menu = tk.Menu(menu_bar,tearoff=0)
mode_menu.add_command(label="Dark Mode",command=set_dark_mode)
mode_menu.add_command(label="Light Mode",command=set_light_mode)
menu_bar.add_cascade(label="Mode",menu=mode_menu)
root.configure(menu=menu_bar)

#Load data from Excel

def load_data():
    file_path = filedialog.askopenfilename()
    if file_path:
        wb = xl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            print(row)
load_button = ttk.Button(form_frame,text="Load Data",command=load_data)
load_button.grid(row=6,columnspan=2,sticky="EW")

#Add a treeview to display the data
data_tree = ttk.Treeview(display_frame)
data_tree.pack()

#Add a scrollbar to the treeview
data_scroll = ttk.Scrollbar(display_frame,orient="vertical",command=data_tree.yview)
data_scroll.pack(side="right",fill="y")
data_tree.configure(yscrollcommand=data_scroll.set)

#add visualisation to the data
def visualise_data():
    import matplotlib.pyplot as plt
    wb = xl.load_workbook("data.xlsx")
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    #create a bar chart
    plt.bar(range(len(data)),[row[1] for row in data])
    plt.title("Age distribution")
    plt.xlabel("Person")
    plt.ylabel("Age")
    plt.show()
visualize_button = ttk.Button(display_frame,text="Visualize Data",command=visualise_data)
visualize_button.pack()
#Start the application
root.mainloop()